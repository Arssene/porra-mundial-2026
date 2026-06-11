"""
Porra Mundial 2026 - Versión Render
  / y /clasificacion -> público
  /admin/*           -> protegido con login
"""
import os, json, unicodedata, secrets, shutil, glob
from zoneinfo import ZoneInfo
from flask import Flask, request, redirect, session, render_template_string, url_for, send_file
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "porristas26")

# ── Rutas de datos persistentes ───────────────────────────────────────────────
BASE_DIR    = os.path.dirname(__file__)
DATA_DIR    = os.path.join(BASE_DIR, "data")
UPLOAD_DIR  = os.path.join(BASE_DIR, "uploads")
STATE_FILE  = os.path.join(DATA_DIR, "state.json")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── Estructura del Excel ──────────────────────────────────────────────────────
PARTIDOS_GRUPOS   = list(range(6, 78))
POSICIONES_GRUPOS = list(range(80, 128))
DIECISEISAVOS_C   = list(range(130, 162))
OCTAVOS_C         = list(range(182, 198))
CUARTOS_C         = list(range(210, 218))
SEMIS_C           = list(range(226, 230))
FINALISTAS_C      = [240, 241]
CAMPEON_C         = 250

BAREMO_DEFAULT = {
    "resultado_exacto": 6, "resultado_un_gol": 4, "resultado_signo": 3,
    "un_gol_falla": 1, "posicion_grupo": 2, "dieciseisavos": 3,
    "octavos": 6, "cuartos": 10, "semifinales": 15, "finalistas": 20, "campeon": 40,
    "bota_oro": 10,
}

# ── Persistencia ──────────────────────────────────────────────────────────────
def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            s = json.load(f)
        # Asegurar claves por si el archivo es de versión anterior
        s.setdefault("jugadores", {})
        s.setdefault("oficiales", {})
        s.setdefault("nombres_partidos", {})
        s.setdefault("nombres_posiciones", {})
        # Rellenar claves de baremo nuevas que no estén en el JSON guardado
        baremo = s.get("baremo", {})
        for k, v in BAREMO_DEFAULT.items():
            baremo.setdefault(k, v)
        s["baremo"] = baremo
        return s
    return {
        "jugadores": {}, "oficiales": {}, "baremo": BAREMO_DEFAULT.copy(),
        "nombres_partidos": {}, "nombres_posiciones": {}
    }

def save_state(s):
    # Backup con timestamp antes de sobreescribir (máximo 3)
    if os.path.exists(STATE_FILE):
        from datetime import datetime
        ts = datetime.now(ZoneInfo("Europe/Madrid")).strftime("%Y%m%d_%H%M")
        shutil.copy2(STATE_FILE, STATE_FILE + f".backup_{ts}")
        # Borrar backups más antiguos si hay más de 3
        backups = sorted(glob.glob(STATE_FILE + ".backup_*"))
        for old in backups[:-3]:
            os.remove(old)
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(s, f, ensure_ascii=False, indent=2)

def list_backups():
    backups = sorted(glob.glob(STATE_FILE + ".backup_*"), reverse=True)
    result = []
    for b in backups:
        ts = b.split(".backup_")[-1]
        try:
            from datetime import datetime
            dt = datetime.strptime(ts, "%Y%m%d_%H%M")
            label = dt.strftime("%d/%m/%Y %H:%M")
        except:
            label = ts
        result.append({"path": b, "ts": ts, "label": label})
    return result

def restore_backup(ts):
    path = STATE_FILE + f".backup_{ts}"
    if os.path.exists(path):
        shutil.copy2(path, STATE_FILE)
        return True
    return False

# ── Lógica Excel ──────────────────────────────────────────────────────────────
def leer_celda(ws, fila):
    val = ws.cell(row=fila, column=3).value
    return str(val).strip() if val is not None else ""

def signo_resultado(g1, g2):
    if g1 > g2: return "1"
    if g1 < g2: return "2"
    return "X"

def parse_resultado(txt):
    if not txt or txt in ("-", "0|-", ""): return None
    t = txt.strip()
    if "·" in t: t = t.split("·")[1]
    if "|" in t: t = t.split("|")[1].strip()
    if "-" in t:
        g = t.split("-")
        try:
            g1, g2 = int(g[0]), int(g[1])
            return (signo_resultado(g1, g2), g1, g2)
        except: return None
    return None

def normalizar(s):
    if not s: return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")

def extraer_jugador(ruta):
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as e:
        return None, str(e)
    if "Pool" not in wb.sheetnames:
        wb.close()
        return None, "No tiene hoja Pool"
    ws = wb["Pool"]

    # Leer solo las filas que necesitamos
    filas_needed = set([5, 253, CAMPEON_C] + PARTIDOS_GRUPOS + POSICIONES_GRUPOS +
                       DIECISEISAVOS_C + OCTAVOS_C + CUARTOS_C + SEMIS_C + FINALISTAS_C)
    celdas = {}
    for row in ws.iter_rows(min_row=1, max_row=max(filas_needed), min_col=2, max_col=3):
        for cell in row:
            if cell.row in filas_needed:
                celdas[(cell.row, cell.column)] = str(cell.value).strip() if cell.value is not None else ""
    wb.close()

    def get(fila, col=3):
        return celdas.get((fila, col), "")

    nombre = get(5)
    if not nombre or nombre in ("Nombre", ""):
        nombre = os.path.splitext(os.path.basename(ruta))[0]
    datos = {
        "nombre": nombre,
        "partidos":      {str(f): get(f) for f in PARTIDOS_GRUPOS},
        "posiciones":    {str(f): get(f) for f in POSICIONES_GRUPOS},
        "dieciseisavos": {str(f): get(f) for f in DIECISEISAVOS_C},
        "octavos":       {str(f): get(f) for f in OCTAVOS_C},
        "cuartos":       {str(f): get(f) for f in CUARTOS_C},
        "semis":         {str(f): get(f) for f in SEMIS_C},
        "finalistas":    {str(f): get(f) for f in FINALISTAS_C},
        "campeon":       get(CAMPEON_C),
        "goleador":      get(253),
    }
    nombres_p   = {str(f): (get(f, 2) or f"Partido {f}") for f in PARTIDOS_GRUPOS}
    nombres_pos = {str(f): (get(f, 2) or f"Posición {f}") for f in POSICIONES_GRUPOS}
    return datos, None, nombres_p, nombres_pos

def calcular_puntos(jugador, oficiales, baremo):
    pts = 0; det = {}
    # Partidos grupo
    p = 0; exactos = 0
    for f, pred in jugador["partidos"].items():
        of = oficiales.get(f"p{f}", "")
        if not of: continue
        r_of = parse_resultado(of); r_p = parse_resultado(pred)
        if not r_of or not r_p: continue
        if r_of[1]==r_p[1] and r_of[2]==r_p[2]:
            p += baremo["resultado_exacto"]; exactos += 1
        elif r_of[0]==r_p[0] and (r_of[1]==r_p[1] or r_of[2]==r_p[2]): p += baremo["resultado_un_gol"]
        elif r_of[0]==r_p[0]: p += baremo["resultado_signo"]
        elif r_of[1]==r_p[1] or r_of[2]==r_p[2]: p += baremo["un_gol_falla"]
    det["partidos"] = p; det["exactos"] = exactos; pts += p
    # Posiciones grupo
    pp = 0
    for f, pred in jugador["posiciones"].items():
        of = oficiales.get(f"pos{f}", "")
        if of and pred and normalizar(of)==normalizar(pred): pp += baremo["posicion_grupo"]
    det["posiciones"] = pp; pts += pp
    # Fases eliminatorias
    def fase_pts(key_jugador, prefix, baremo_key):
        p2 = 0
        of_set = set(normalizar(v) for k,v in oficiales.items() if k.startswith(prefix) and v)
        for pred in jugador[key_jugador].values():
            if normalizar(pred) in of_set: p2 += baremo[baremo_key]
        return p2
    det["dieciseisavos"] = fase_pts("dieciseisavos","d16_","dieciseisavos"); pts += det["dieciseisavos"]
    det["octavos"]       = fase_pts("octavos","oct_","octavos");             pts += det["octavos"]
    det["cuartos"]       = fase_pts("cuartos","cua_","cuartos");             pts += det["cuartos"]
    det["semis"]         = fase_pts("semis","sem_","semifinales");           pts += det["semis"]
    det["finalistas"]    = fase_pts("finalistas","fin_","finalistas");       pts += det["finalistas"]
    # Campeón
    of_cam = normalizar(oficiales.get("campeon",""))
    cam = baremo["campeon"] if of_cam and normalizar(jugador["campeon"])==of_cam else 0
    det["campeon"] = cam; pts += cam
    # Bota de Oro
    of_gol = normalizar(oficiales.get("bota_oro",""))
    gol = baremo["bota_oro"] if of_gol and normalizar(jugador.get("goleador",""))==of_gol else 0
    det["bota_oro"] = gol; pts += gol
    det["total"] = pts
    return det

# ── CSS y HTML base ───────────────────────────────────────────────────────────
CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', Arial, sans-serif; background: #0a1628; color: #e0e0e0; min-height: 100vh; }
header { background: linear-gradient(135deg, #1a3a6e, #c8102e); padding: 14px 20px; display: flex; align-items: center; gap: 12px; }
header h1 { font-size: 1.3em; color: white; }
nav { background: #0d1f3c; display: flex; gap: 2px; padding: 0 10px; border-bottom: 2px solid #c8102e; flex-wrap: wrap; }
nav a { padding: 10px 14px; color: #aaa; text-decoration: none; font-weight: 600; font-size: 0.83em; white-space: nowrap; }
nav a:hover, nav a.active { color: white; background: #1a3a6e; }
main { max-width: 1100px; margin: 20px auto; padding: 0 12px; }
.card { background: #0d1f3c; border: 1px solid #1a3a6e; border-radius: 10px; padding: 18px; margin-bottom: 18px; }
.card h2 { color: #4a9eff; margin-bottom: 14px; font-size: 1.02em; }
.btn { padding: 9px 18px; border: none; border-radius: 6px; cursor: pointer; font-weight: 600; font-size: 0.88em; text-decoration: none; display: inline-block; }
.btn-red { background: #c8102e; color: white; } .btn-red:hover { background: #a00d24; }
.btn-green { background: #1a7a3e; color: white; } .btn-green:hover { background: #145f30; }
.btn-blue { background: #1a3a6e; color: white; }
input[type=text], input[type=password], input[type=number] { background: #0a1628; border: 1px solid #1a3a6e; color: #e0e0e0; padding: 7px 10px; border-radius: 5px; }
input[type=file] { color: #aaa; }
label { display: block; margin-bottom: 4px; color: #aaa; font-size: 0.85em; }
.form-row { margin-bottom: 10px; }
.grid-2 { display: grid; grid-template-columns: 1fr 180px; gap: 10px; align-items: center; }
.tabla-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
.tabla { width: 100%; border-collapse: collapse; font-size: 0.82em; min-width: 540px; }
.tabla th { background: #1a3a6e; padding: 9px 5px; text-align: center; color: #4a9eff; font-size: 0.8em; white-space: nowrap; }
.tabla td { padding: 8px 5px; border-bottom: 1px solid #1a3a6e; text-align: center; white-space: nowrap; }
.tabla tr:hover td { background: #111f3a; }
.nombre { text-align: left !important; font-weight: 600; color: white; white-space: normal !important; min-width: 90px; }
.total { font-weight: 700; color: #ffd700; font-size: 1.05em; }
.exactos { color: #4adf7a; font-weight: 600; }
.pos-medal { font-size: 1.05em; }
.msg-ok { padding: 10px 15px; border-radius: 6px; margin-bottom: 14px; background: #1a3a1a; border: 1px solid #1a7a3e; color: #4adf7a; }
.msg-err { padding: 10px 15px; border-radius: 6px; margin-bottom: 14px; background: #3a1a1a; border: 1px solid #7a1a1a; color: #df4a4a; }
.sep { border: none; border-top: 1px solid #1a3a6e; margin: 18px 0; }
.baremo-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; }
.baremo-item label { font-size: 0.8em; }
.baremo-item input { width: 70px; }
.tag { background: #1a3a6e; padding: 4px 11px; border-radius: 20px; font-size: 0.82em; display: inline-block; margin: 3px; }
details summary { cursor: pointer; color: #4a9eff; font-weight: 600; padding: 8px 0; user-select: none; }
.login-box { max-width: 360px; margin: 60px auto; }
.tendencia { font-size: 1.1em; letter-spacing: 1px; white-space: nowrap; }
@media (max-width: 500px) {
  header h1 { font-size: 1.05em; }
  nav a { padding: 8px 9px; font-size: 0.76em; }
  .card { padding: 12px; }
  .grid-2 { grid-template-columns: 1fr; }
}
"""

def base(content, tab="pub", admin=False):
    nav_pub = (
        f'<a href="/" class="{"active" if tab=="pub" else ""}">🏆 Clasificación</a>'
        f'<a href="/estadisticas" class="{"active" if tab=="stats" else ""}">📊 Estadísticas</a>'
        f'<a href="/baremo" class="{"active" if tab=="baremo_pub" else ""}">📋 Baremo</a>'
    )
    nav_admin = ""
    if admin:
        nav_admin = f'''
        <a href="/admin" class="{"active" if tab=="home" else ""}">🏠 Admin</a>
        <a href="/admin/cargar" class="{"active" if tab=="cargar" else ""}">📂 Cargar Porras</a>
        <a href="/admin/resultados" class="{"active" if tab=="resultados" else ""}">📋 Resultados</a>
        <a href="/admin/baremo" class="{"active" if tab=="baremo" else ""}">⚙️ Baremo</a>
        <a href="/admin/goleadores" class="{"active" if tab=="goleadores" else ""}">👟 Goleadores</a>
        <a href="/admin/logout" style="margin-left:auto;color:#c8102e">Salir</a>'''
    else:
        nav_admin = '<a href="/admin" style="margin-left:auto;color:#4a9eff;font-size:0.8em">Admin</a>'
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>🏆 Porra Mundial 2026</title>
<style>{CSS}</style></head>
<body>
<header><span style="font-size:2em">⚽</span><h1>Porra Mundial 2026</h1></header>
<nav>{nav_pub}{nav_admin}</nav>
<main>{content}</main>
</body></html>"""

# ── Auth ──────────────────────────────────────────────────────────────────────
def is_admin():
    return session.get("admin") is True

def require_admin():
    if not is_admin():
        return redirect("/admin/login")
    return None


def tendencia_jugador(nombre, historial):
    """Devuelve HTML con flechas de las últimas 3 actualizaciones."""
    if not historial:
        return ""
    flechas = []
    snapshots = historial[-3:]
    for i in range(1, len(snapshots)):
        prev = snapshots[i-1]["ranking"].get(nombre)
        curr = snapshots[i]["ranking"].get(nombre)
        if prev is None or curr is None:
            flechas.append('<span style="color:#aaa">➖</span>')
        elif curr < prev:
            diff = prev - curr
            flechas.append(f'<span style="color:#4adf7a">↑{diff}</span>')
        elif curr > prev:
            diff = curr - prev
            flechas.append(f'<span style="color:#df4a4a">↓{diff}</span>')
        else:
            flechas.append('<span style="color:#aaa">➖</span>')
    return " ".join(flechas)


def render_backups():
    bks = list_backups()
    if not bks:
        return "<p style='color:#aaa;font-size:0.85em'>⚠️ Aún no hay backups (se crearán al guardar por primera vez)</p>"
    html = ""
    for b in bks:
        label = b["label"]
        ts = b["ts"]
        confirm_msg = f"¿Restaurar backup del {label}?"
        html += (
            f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:8px'>"
            f"<span style='color:#4adf7a;font-size:0.85em'>✅ {label}</span>"
            f"<form method='post' action='/admin/restaurar_backup' style='margin:0' "
            f"onsubmit='return confirm(\"{confirm_msg}\")'>"
            f"<input type='hidden' name='ts' value='{ts}'>"
            f"<button class='btn' style='background:#7a4a00;color:white;padding:4px 10px;font-size:0.8em'"
            f" type='submit'>⏪ Restaurar</button>"
            f"</form>"" <a href='/admin/descargar_json' class='btn btn-blue' style='margin-left:8px'>📥 Descargar JSON</a>""</div>"
        )
    return html

# ── Rutas públicas ────────────────────────────────────────────────────────────
@app.route("/")
def clasificacion():
    s = load_state()
    if not s["jugadores"]:
        content = '<div class="card"><h2>🏆 Clasificación</h2><p style="color:#aaa">Todavía no hay participantes cargados.</p></div>'
        return base(content, "pub")

    resultados = []
    for nombre, datos in s["jugadores"].items():
        det = calcular_puntos(datos, s["oficiales"], s["baremo"])
        resultados.append((nombre, det, datos))
    resultados.sort(key=lambda x: -x[1]["total"])

    medallas = ["🥇","🥈","🥉"]
    historial = s.get("historial_rankings", [])
    rows = ""
    for i, (nombre, det, datos) in enumerate(resultados):
        medal = medallas[i] if i < 3 else f"{i+1}º"
        tend = tendencia_jugador(nombre, historial)
        rows += f"""<tr>
            <td class="pos-medal">{medal}</td>
            <td class="nombre">{nombre}</td>
            <td class="total">{det['total']}</td>
            <td class="tendencia">{tend}</td>
            <td>{det['partidos']}</td>
            <td class="exactos">{det['exactos']}</td>
            <td>{det['posiciones']}</td>
            <td>{det['dieciseisavos']}</td><td>{det['octavos']}</td>
            <td>{det['cuartos']}</td><td>{det['semis']}</td>
            <td>{det['finalistas']}</td><td>{det['campeon']}</td>
            <td style='font-size:0.78em;color:#ffd700'>{datos.get('campeon','') or '—'}</td>
            <td style='font-size:0.78em;color:#aaa'>{_fin[0] if (_fin:=list(datos.get('finalistas',{}).values())) else '—'}</td>
            <td style='font-size:0.78em;color:#aaa'>{_fin[1] if len(_fin)>1 else '—'}</td>
            <td style='font-size:0.78em;color:#4adf7a'>{datos.get('goleador','') or '—'}</td>
        </tr>"""

    content = f"""<div class="card">
        <h2>🏆 Clasificación General — {len(resultados)} participantes</h2>
        {"<p style=\'color:#aaa;font-size:0.82em;margin-bottom:12px\'>⏱ Última actualización: " + s.get("ultima_actualizacion","—") + "</p>" if s.get("ultima_actualizacion") else ""}
        <div class="tabla-wrap">
        <table class="tabla">
            <thead><tr>
                <th>#</th><th style="text-align:left">Jugador</th>
                <th>TOTAL</th><th>Tend.</th>
                <th>Pts<br>Grupos</th><th>🎯<br>Exactos</th><th>Clasif.<br>Grupos</th>
                <th>1/16</th><th>1/8</th><th>1/4</th><th>1/2</th>
                <th>Final</th><th>Campeón</th><th style="font-size:0.75em">Campeón</th><th style="font-size:0.75em">Finalista</th><th style="font-size:0.75em">Finalista</th><th style="font-size:0.75em">Goleador</th>
            </tr></thead>
            <tbody>{rows}</tbody>
        </table>
        </div>
    </div>"""
    return base(content, "pub")


@app.route("/baremo")
def baremo_publico():
    s = load_state()
    b = s["baremo"]
    items = [
        ("Resultado exacto (ej: 2-1 ✓ 2-1)", b["resultado_exacto"]),
        ("Resultado + un gol (ej: 2-1 ✓ 2-0)", b["resultado_un_gol"]),
        ("Signo acertado (ej: 2-1 ✓ 1-0)", b["resultado_signo"]),
        ("Falla resultado, acierta un gol", b["un_gol_falla"]),
        ("Posición en grupo acertada", b["posicion_grupo"]),
        ("Clasificado a dieciseisavos", b["dieciseisavos"]),
        ("Clasificado a octavos", b["octavos"]),
        ("Clasificado a cuartos", b["cuartos"]),
        ("Semifinalista", b["semifinales"]),
        ("Finalista", b["finalistas"]),
        ("Campeón", b["campeon"]),
        ("👟 Bota de Oro (máximo goleador)", b["bota_oro"]),
    ]
    filas = ""
    for desc, pts in items:
        filas += f"""<tr>
            <td style="text-align:left;color:#ddd">{desc}</td>
            <td style="color:#ffd700;font-weight:700;font-size:1.1em">{pts} pts</td>
        </tr>"""
    content = f"""<div class="card">
        <h2>📋 Baremo de puntuación</h2>
        <div class="tabla-wrap">
        <table class="tabla" style="min-width:300px">
            <thead><tr>
                <th style="text-align:left">Concepto</th>
                <th>Puntos</th>
            </tr></thead>
            <tbody>{filas}</tbody>
        </table>
        </div>
    </div>"""
    return base(content, "baremo_pub")


@app.route("/estadisticas")
def estadisticas():
    s = load_state()
    jugadores = s["jugadores"]
    np = s.get("nombres_partidos", {})

    if not jugadores:
        content = '<div class="card"><h2>📊 Estadísticas</h2><p style="color:#aaa">Todavía no hay participantes cargados.</p></div>'
        return base(content, "stats")

    # --- Campeones votados ---
    from collections import Counter
    camp_count = Counter()
    for j in jugadores.values():
        c = normalizar(j.get("campeon", ""))
        if c and c not in ("WF", ""):
            camp_count[c] += 1

    # --- Goleadores votados ---
    gol_count = Counter()
    for j in jugadores.values():
        g = j.get("goleador", "").strip()
        if g and g not in ("Escribe un jugador", "") and normalizar(g) != "ESCRIBE UN JUGADOR":
            # Use title case for display
            gol_count[g.strip().title()] += 1

    # --- Partido más acertado (signo) ---
    partido_aciertos = {}
    for f in PARTIDOS_GRUPOS:
        of = s["oficiales"].get(f"p{f}", "")
        if not of: continue
        r_of = parse_resultado(of)
        if not r_of: continue
        aciertos = 0; total = 0
        for j in jugadores.values():
            pred = j["partidos"].get(str(f), "")
            r_p = parse_resultado(pred)
            if r_p:
                total += 1
                if r_of[0] == r_p[0]: aciertos += 1
        if total > 0:
            partido_aciertos[f] = {"aciertos": aciertos, "total": total, "pct": round(aciertos/total*100)}

    mas_acertado = sorted(partido_aciertos.items(), key=lambda x: -x[1]["pct"])[:5]

    # --- Resultado exacto más votado ---
    exactos_votados = Counter()
    for j in jugadores.values():
        for f, pred in j["partidos"].items():
            r_p = parse_resultado(pred)
            if not r_p: continue
            of = s["oficiales"].get(f"p{f}", "")
            r_of = parse_resultado(of)
            if r_of and r_of[1]==r_p[1] and r_of[2]==r_p[2]:
                nombre_p = np.get(str(f), f"Partido {f}")
                exactos_votados[f"{nombre_p} ({pred})"] += 1

    top_exactos = exactos_votados.most_common(5)

    # --- Generar colores ---
    COLORS = ["#4a9eff","#c8102e","#4adf7a","#ffd700","#df4a9a","#df944a",
              "#9a4adf","#4adfdf","#df4a4a","#a0df4a","#4a6eff","#ff9a4a"]

    def pie_chart(counter, title, svg_id):
        import math
        if not counter:
            return "<div class='card'><h2>" + title + "</h2><p style='color:#aaa'>Sin datos todavía</p></div>"
        total = sum(counter.values())
        items = counter.most_common()
        cx, cy, r = 120, 120, 100
        angle = -math.pi/2
        paths = ""
        legend = ""
        for i, (label, val) in enumerate(items):
            frac = val / total
            a2 = angle + frac * 2 * math.pi
            x1 = cx + r * math.cos(angle)
            y1 = cy + r * math.sin(angle)
            x2 = cx + r * math.cos(a2)
            y2 = cy + r * math.sin(a2)
            large = 1 if frac > 0.5 else 0
            color = COLORS[i % len(COLORS)]
            paths += "<path d='M" + str(cx) + "," + str(cy) + " L" + str(round(x1,1)) + "," + str(round(y1,1)) + " A" + str(r) + "," + str(r) + " 0 " + str(large) + ",1 " + str(round(x2,1)) + "," + str(round(y2,1)) + " Z' fill='" + color + "' stroke='#0a1628' stroke-width='2'/>"
            pct = round(frac*100)
            legend += "<div style='display:flex;align-items:center;gap:6px;margin-bottom:5px'><div style='width:12px;height:12px;border-radius:2px;background:" + color + ";flex-shrink:0'></div><span style='font-size:0.82em;color:#ddd'>" + label + " <b style='color:#ffd700'>(" + str(val) + ")</b> " + str(pct) + "%</span></div>"
            angle = a2
        svg = "<svg viewBox='0 0 240 240' width='200' height='200'>" + paths + "<circle cx='" + str(cx) + "' cy='" + str(cy) + "' r='40' fill='#0d1f3c'/><text x='" + str(cx) + "' y='" + str(cy) + "' text-anchor='middle' dy='5' fill='#fff' font-size='13' font-weight='bold'>" + str(total) + "</text></svg>"
        return "<div class='card'><h2>" + title + "</h2><div style='display:flex;flex-wrap:wrap;gap:20px;align-items:flex-start'><div>" + svg + "</div><div style='flex:1;min-width:160px'>" + legend + "</div></div></div>"

    # Partidos más acertados
    if mas_acertado:
        filas_pa = ""
        for f, d in mas_acertado:
            nombre_p = np.get(str(f), "Partido " + str(f))
            bw = str(d["pct"])
            filas_pa += (
                "<div style='margin-bottom:12px'>"
                "<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                "<span style='color:#ddd;font-size:0.85em'>" + nombre_p + "</span>"
                "<span style='color:#ffd700;font-weight:600'>" + str(d["aciertos"]) + "/" + str(d["total"]) + " (" + str(d["pct"]) + "%)</span>"
                "</div>"
                "<div style='background:#1a3a6e;border-radius:4px;height:10px'>"
                "<div style='background:#4a9eff;border-radius:4px;height:10px;width:" + bw + "%'></div>"
                "</div></div>"
            )
        card_partidos = "<div class='card'><h2>⚽ Partidos más acertados (signo)</h2>" + filas_pa + "</div>"
    else:
        card_partidos = "<div class='card'><h2>⚽ Partidos más acertados</h2><p style='color:#aaa'>Sin resultados oficiales todavía</p></div>"

    # Exactos más votados
    if top_exactos:
        filas_ex = ""
        max_ex = top_exactos[0][1] if top_exactos else 1
        for label, cnt in top_exactos:
            bw = str(round(cnt/max_ex*100))
            filas_ex += (
                "<div style='margin-bottom:12px'>"
                "<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                "<span style='color:#ddd;font-size:0.85em'>" + label + "</span>"
                "<span style='color:#4adf7a;font-weight:600'>" + str(cnt) + " aciertos</span>"
                "</div>"
                "<div style='background:#1a3a6e;border-radius:4px;height:10px'>"
                "<div style='background:#4adf7a;border-radius:4px;height:10px;width:" + bw + "%'></div>"
                "</div></div>"
            )
        card_exactos = "<div class='card'><h2>🎯 Resultados exactos más acertados</h2>" + filas_ex + "</div>"
    else:
        card_exactos = "<div class='card'><h2>🎯 Resultados exactos más acertados</h2><p style='color:#aaa'>Sin exactos todavía</p></div>"

    content = (
        pie_chart(camp_count, "🏆 Campeones votados", "pie_camp") +
        pie_chart(gol_count, "👟 Goleadores votados", "pie_gol") +
        card_partidos +
        card_exactos
    )
    return base(content, "stats")

# ── Rutas admin ───────────────────────────────────────────────────────────────
@app.route("/admin/login", methods=["GET","POST"])
def login():
    error = ""
    if request.method == "POST":
        if request.form.get("user")==ADMIN_USER and request.form.get("pass")==ADMIN_PASS:
            session["admin"] = True
            return redirect("/admin")
        error = "Usuario o contraseña incorrectos"
    content = f"""<div class="card login-box">
        <h2>🔐 Acceso Admin</h2>
        {"<div class='msg-err'>"+error+"</div>" if error else ""}
        <form method="post">
            <div class="form-row"><label>Usuario</label><input type="text" name="user" style="width:100%"></div>
            <div class="form-row"><label>Contraseña</label><input type="password" name="pass" style="width:100%"></div>
            <br><button class="btn btn-red" type="submit">Entrar</button>
        </form>
    </div>"""
    return base(content, "login")

@app.route("/admin/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/admin")
def admin_home():
    r = require_admin()
    if r: return r
    s = load_state()
    n_j = len(s["jugadores"])
    n_of = sum(1 for v in s["oficiales"].values() if v)
    content = f"""<div class="card">
        <h2>🏠 Panel de Administración</h2>
        <p style="color:#aaa;margin-bottom:16px">Bienvenido, <b>{ADMIN_USER}</b></p>
        <p>👥 Participantes cargados: <b style="color:#ffd700">{n_j}</b></p>
        <p style="margin-top:8px">📋 Resultados introducidos: <b style="color:#4adf7a">{n_of}</b></p>
        <hr class="sep">
        <a href="/admin/cargar" class="btn btn-blue" style="margin-right:8px">📂 Cargar Porras</a>
        <a href="/admin/resultados" class="btn btn-blue" style="margin-right:8px">📋 Resultados</a>
        <a href="/admin/baremo" class="btn btn-blue">⚙️ Baremo</a>
        <hr class="sep">
        <h2>🛡 Copia de seguridad</h2>
        {"<div class='msg-ok'>✅ Backup restaurado correctamente</div>" if request.args.get("msg")=="backup_ok" else ""}
        {"<div class='msg-err'>⚠️ No hay backup disponible</div>" if request.args.get("msg")=="backup_error" else ""}
        <p style="color:#aaa;font-size:0.88em;margin-bottom:12px">
            Cada vez que se guardan resultados, baremo o jugadores se crea automáticamente una copia de seguridad del estado anterior.
            Si algo sale mal puedes restaurarla aquí.
        </p>
{render_backups()}
    </div>"""
    return base(content, "home", admin=True)

@app.route("/admin/cargar", methods=["GET","POST"])
def admin_cargar():
    r = require_admin()
    if r: return r
    s = load_state()
    msg = ""; msg_class = "msg-ok"

    if request.method == "POST":
        if "limpiar" in request.form:
            s["jugadores"].clear()
            s["nombres_partidos"] = {}
            s["nombres_posiciones"] = {}
            save_state(s)
            msg = "✅ Todos los jugadores eliminados"
        elif "json_file" in request.files and request.files["json_file"].filename:
            jf = request.files["json_file"]
            try:
                nuevo = json.load(jf.stream)
                # Validar estructura mínima
                if "jugadores" not in nuevo:
                    raise ValueError("El JSON no tiene la clave 'jugadores'")
                # Rellenar claves de baremo que falten
                baremo = nuevo.get("baremo", {})
                for k, v in BAREMO_DEFAULT.items():
                    baremo.setdefault(k, v)
                nuevo["baremo"] = baremo
                nuevo.setdefault("oficiales", {})
                nuevo.setdefault("nombres_partidos", {})
                nuevo.setdefault("nombres_posiciones", {})
                save_state(nuevo)
                s = nuevo
                msg = f"✅ JSON cargado correctamente: {len(nuevo['jugadores'])} jugadores"
            except Exception as e:
                msg = f"⚠️ Error al leer el JSON: {e}"
                msg_class = "msg-err"
        else:
            archivos = request.files.getlist("archivos")
            cargados = []; errores = []
            for f in archivos:
                if not f.filename.endswith(".xlsx"): continue
                ruta = os.path.join(UPLOAD_DIR, f.filename)
                f.save(ruta)
                resultado = extraer_jugador(ruta)
                if resultado[1]:
                    errores.append(f"{f.filename}: {resultado[1]}")
                else:
                    datos, _, np, npos = resultado
                    s["jugadores"][datos["nombre"]] = datos
                    if not s["nombres_partidos"]:
                        s["nombres_partidos"] = np
                        s["nombres_posiciones"] = npos
                    cargados.append(datos["nombre"])
            save_state(s)
            if cargados:
                msg = f"✅ Cargados: {', '.join(cargados)}"
                if errores: msg += f" | ⚠️ {', '.join(errores)}"
            else:
                msg = f"⚠️ {'; '.join(errores)}"; msg_class = "msg-err"

    from urllib.parse import quote
    filas_j = ""
    for nombre in s["jugadores"]:
        enc = quote(nombre)
        safe = nombre.replace("'", "\\'")
        filas_j += (
            "<tr>"
            "<td style='text-align:left;color:white;font-weight:600'>" + nombre + "</td>"
            "<td style='white-space:nowrap'>"
            "<a href='/admin/jugador/" + enc + "' class='btn btn-blue' style='padding:4px 10px;font-size:0.8em;margin-right:4px'>👁 Ver</a> "
            "<a href='/admin/renombrar/" + enc + "' class='btn btn-blue' style='padding:4px 10px;font-size:0.8em;margin-right:4px'>✏️ Renombrar</a> "
            "<a href='/admin/eliminar/" + enc + "' class='btn' style='padding:4px 10px;font-size:0.8em;background:#7a1a1a;color:white' "
            "onclick='return confirm(\"Eliminar a " + safe + "?\")'>🗑 Eliminar</a>"
            "</td></tr>"
        )
    tabla_j = (
        "<div class='tabla-wrap'><table class='tabla' style='min-width:300px'>"
        "<thead><tr><th style='text-align:left'>Jugador</th><th>Acciones</th></tr></thead>"
        "<tbody>" + filas_j + "</tbody></table></div>"
    ) if filas_j else "<p style='color:#666'>Ninguno todavía</p>"

    n_j = len(s["jugadores"])
    msg_html = "<div class='" + msg_class + "'>" + msg + "</div>" if msg else ""
    content = (
        "<div class='card'>"
        "<h2>📂 Cargar archivos de porra</h2>"
        + msg_html +
        "<form method='post' enctype='multipart/form-data'>"
        "<div class='form-row'>"
        "<label>Selecciona uno o varios archivos Excel (.xlsx)</label>"
        "<input type='file' name='archivos' multiple accept='.xlsx'>"
        "</div>"
        "<button class='btn btn-red' type='submit'>⬆ Cargar</button>"
        "</form>"
        "<hr class='sep'>"
        "<details><summary>▶ Subir JSON completo (avanzado)</summary>"
        "<div style='margin-top:10px'>"
        "<p style='color:#aaa;font-size:0.85em;margin-bottom:10px'>"
        "Sube un archivo state.json generado en local. <b>Esto reemplaza todos los datos actuales</b> "
        "(jugadores, resultados, baremo e historial)."
        "</p>"
        "<form method='post' enctype='multipart/form-data' "
        "onsubmit=\"return confirm('Esto reemplazará TODOS los datos actuales por los del JSON. ¿Continuar?')\">"
        "<div class='form-row'>"
        "<input type='file' name='json_file' accept='.json'>"
        "</div>"
        "<button class='btn btn-green' type='submit'>📤 Subir JSON</button>"
        "</form>"
        "</div></details>"
        "<hr class='sep'>"
        "<h2>Participantes (" + str(n_j) + ")</h2>"
        "<div style='margin-bottom:16px'>" + tabla_j + "</div>"
        "<form method='post'>"
        "<button class='btn btn-blue' name='limpiar' value='1' "
        "onclick='return confirm(\"¿Borrar todos los jugadores?\")'>🗑 Limpiar todo</button>"
        "</form></div>"
    )
    return base(content, "cargar", admin=True)

@app.route("/admin/descargar_json")
def admin_descargar_json():
    r = require_admin()
    if r: return r
    if not os.path.exists(STATE_FILE):
        return redirect("/admin/cargar")
    return send_file(STATE_FILE, as_attachment=True, download_name="state.json")


@app.route("/admin/resultados", methods=["GET","POST"])
def admin_resultados():
    r = require_admin()
    if r: return r
    s = load_state()
    msg = ""
    if request.method == "POST":
        for k, v in request.form.items():
            s["oficiales"][k] = v.strip()
        from datetime import datetime
        s["ultima_actualizacion"] = datetime.now(ZoneInfo("Europe/Madrid")).strftime("%d/%m/%Y %H:%M")
        # Guardar snapshot del ranking actual antes de recalcular
        if s["jugadores"]:
            ranking_actual = {}
            resultados_snap = []
            for nombre, datos in s["jugadores"].items():
                det = calcular_puntos(datos, s["oficiales"], s["baremo"])
                resultados_snap.append((nombre, det["total"]))
            resultados_snap.sort(key=lambda x: -x[1])
            for pos, (nombre, _) in enumerate(resultados_snap, start=1):
                ranking_actual[nombre] = pos
            historial = s.get("historial_rankings", [])
            historial.append({"fecha": s["ultima_actualizacion"], "ranking": ranking_actual})
            s["historial_rankings"] = historial[-3:]  # solo últimas 3
        save_state(s)
        msg = "✅ Resultados guardados"

    np = s.get("nombres_partidos", {})
    npos = s.get("nombres_posiciones", {})

    partidos_html = ""
    for f in PARTIDOS_GRUPOS:
        nombre = np.get(str(f), f"Partido {f}")
        val = s["oficiales"].get(f"p{f}", "")
        partidos_html += f"""<div class="form-row grid-2">
            <label style="color:#ddd;font-size:0.88em">{nombre}</label>
            <input type="text" name="p{f}" value="{val}" placeholder="2-1" style="width:120px">
        </div>"""

    pos_html = ""
    for f in POSICIONES_GRUPOS:
        nombre = npos.get(str(f), f"Posición {f}")
        val = s["oficiales"].get(f"pos{f}", "")
        pos_html += f"""<div class="form-row grid-2">
            <label style="color:#ddd;font-size:0.88em">{nombre}</label>
            <input type="text" name="pos{f}" value="{val}" placeholder="Equipo" style="width:160px">
        </div>"""

    def fase_inputs(filas, prefix, label):
        h = f"<h3 style='color:#4a9eff;margin:14px 0 8px'>{label}</h3>"
        for f in filas:
            val = s["oficiales"].get(f"{prefix}{f}", "")
            h += f"""<div class="form-row grid-2">
                <label style="color:#aaa;font-size:0.83em">Clasificado</label>
                <input type="text" name="{prefix}{f}" value="{val}" placeholder="Equipo" style="width:160px">
            </div>"""
        return h

    campeon_val = s["oficiales"].get("campeon", "")
    content = f"""<div class="card">
        <h2>📋 Resultados Oficiales</h2>
        {"<div class='msg-ok'>"+msg+"</div>" if msg else ""}
        <form method="post">
        <details open><summary>▶ Partidos Fase de Grupos</summary>
            <div style="margin-top:8px">{partidos_html}</div>
        </details>
        <hr class="sep">
        <details><summary>▶ Clasificaciones de Grupos</summary>
            <div style="margin-top:8px">{pos_html}</div>
        </details>
        <hr class="sep">
        <details><summary>▶ Eliminatorias</summary>
            <div style="margin-top:8px">
            {fase_inputs(DIECISEISAVOS_C,"d16_","Clasificados a Dieciseisavos")}
            {fase_inputs(OCTAVOS_C,"oct_","Clasificados a Octavos")}
            {fase_inputs(CUARTOS_C,"cua_","Clasificados a Cuartos")}
            {fase_inputs(SEMIS_C,"sem_","Semifinalistas")}
            {fase_inputs(FINALISTAS_C,"fin_","Finalistas")}
            <h3 style='color:#4a9eff;margin:14px 0 8px'>🥇 Campeón y Bota de Oro</h3>
            <div class="form-row grid-2">
                <label style="color:#ddd">Campeón del Mundial</label>
                <input type="text" name="campeon" value="{campeon_val}" placeholder="Equipo" style="width:160px">
            </div>
            <div class="form-row grid-2">
                <label style="color:#ddd">👟 Bota de Oro (máximo goleador)</label>
                <input type="text" name="bota_oro" value="{s['oficiales'].get('bota_oro','')}" placeholder="Nombre jugador" style="width:160px">
            </div>
            </div>
        </details>
        <hr class="sep">
        <button class="btn btn-green" type="submit">💾 Guardar resultados</button>
        </form>
    </div>"""
    return base(content, "resultados", admin=True)

@app.route("/admin/goleadores", methods=["GET","POST"])
def admin_goleadores():
    r = require_admin()
    if r: return r
    s = load_state()
    msg = ""

    if request.method == "POST":
        for nombre, jugador in s["jugadores"].items():
            key = "gol_" + nombre
            if key in request.form:
                nuevo = request.form[key].strip()
                jugador["goleador"] = nuevo
        save_state(s)
        msg = "✅ Goleadores guardados"

    # Build table sorted by goleador name for easy review
    jugadores_sorted = sorted(
        s["jugadores"].items(),
        key=lambda x: (x[1].get("goleador","") or "").upper()
    )

    filas = ""
    for nombre, jugador in jugadores_sorted:
        gol = jugador.get("goleador", "") or ""
        key = "gol_" + nombre
        color = "#aaa" if not gol or gol == "Escribe un jugador" else "#4adf7a"
        filas += (
            "<tr>"
            "<td style='text-align:left;color:white;font-size:0.88em'>" + nombre + "</td>"
            "<td><input type='text' name='" + key + "' value='" + gol.replace("'", "&#39;") + "' "
            "style='width:200px;padding:4px 8px;color:" + color + "'></td>"
            "</tr>"
        )

    msg_html = "<div class='msg-ok'>" + msg + "</div>" if msg else ""
    content = (
        "<div class='card'>"
        "<h2>👟 Editar Goleadores</h2>"
        "<p style='color:#aaa;font-size:0.85em;margin-bottom:14px'>"
        "Ordenados por goleador para que puedas agrupar y normalizar fácilmente. "
        "Edita los que necesites y guarda todo con un clic.</p>"
        + msg_html +
        "<form method='post'>"
        "<div class='tabla-wrap'>"
        "<table class='tabla' style='min-width:350px'>"
        "<thead><tr>"
        "<th style='text-align:left'>Jugador</th>"
        "<th style='text-align:left'>Goleador</th>"
        "</tr></thead>"
        "<tbody>" + filas + "</tbody>"
        "</table></div>"
        "<br>"
        "<button class='btn btn-green' type='submit'>💾 Guardar todos</button>"
        "</form>"
        "</div>"
    )
    return base(content, "goleadores", admin=True)


@app.route("/admin/baremo", methods=["GET","POST"])
def admin_baremo():
    r = require_admin()
    if r: return r
    s = load_state()
    msg = ""
    if request.method == "POST":
        for k in BAREMO_DEFAULT:
            try: s["baremo"][k] = int(request.form.get(k, s["baremo"][k]))
            except: pass
        save_state(s)
        msg = "✅ Baremo guardado"

    items = [
        ("resultado_exacto","Resultado exacto (2-1 ✓ 2-1)"),
        ("resultado_un_gol","Resultado + un gol (2-1 ✓ 2-0)"),
        ("resultado_signo","Signo acertado (2-1 ✓ 1-0)"),
        ("un_gol_falla","Falla resultado, acierta un gol"),
        ("posicion_grupo","Posición en grupo acertada"),
        ("dieciseisavos","Clasificado a dieciseisavos"),
        ("octavos","Clasificado a octavos"),
        ("cuartos","Clasificado a cuartos"),
        ("semifinales","Semifinalista"),
        ("finalistas","Finalista"),
        ("campeon","Campeón"),
        ("bota_oro","Bota de Oro (máximo goleador)"),
    ]
    inputs = "".join(f"""<div class="baremo-item">
        <label>{desc}</label>
        <input type="number" name="{key}" value="{s['baremo'][key]}" min="0" max="999">
    </div>""" for key, desc in items)
    content = f"""<div class="card">
        <h2>⚙️ Baremo de puntuación</h2>
        {"<div class='msg-ok'>"+msg+"</div>" if msg else ""}
        <form method="post">
        <div class="baremo-grid">{inputs}</div>
        <hr class="sep">
        <button class="btn btn-green" type="submit">💾 Guardar baremo</button>
        </form>
    </div>"""
    return base(content, "baremo", admin=True)

@app.route("/admin/restaurar_backup", methods=["POST"])
def admin_restaurar_backup():
    r = require_admin()
    if r: return r
    ts = request.form.get("ts", "")
    if restore_backup(ts):
        return redirect("/admin?msg=backup_ok")
    return redirect("/admin?msg=backup_error")


@app.route("/admin/eliminar/<nombre>")
def admin_eliminar(nombre):
    r = require_admin()
    if r: return r
    s = load_state()
    if nombre in s["jugadores"]:
        del s["jugadores"][nombre]
        # Limpiar del historial también
        for snap in s.get("historial_rankings", []):
            snap["ranking"].pop(nombre, None)
        save_state(s)
    return redirect("/admin/cargar")

@app.route("/admin/renombrar/<nombre>", methods=["GET","POST"])
def admin_renombrar(nombre):
    r = require_admin()
    if r: return r
    s = load_state()
    msg = ""
    if request.method == "POST":
        nuevo = request.form.get("nuevo_nombre", "").strip()
        if nuevo and nuevo != nombre and nombre in s["jugadores"]:
            s["jugadores"][nuevo] = s["jugadores"].pop(nombre)
            s["jugadores"][nuevo]["nombre"] = nuevo
            for snap in s.get("historial_rankings", []):
                if nombre in snap["ranking"]:
                    snap["ranking"][nuevo] = snap["ranking"].pop(nombre)
            save_state(s)
            return redirect("/admin/cargar")
        msg = "⚠️ Nombre no válido o sin cambios"
    content = f"""<div class="card" style="max-width:420px;margin:40px auto">
        <h2>✏️ Renombrar jugador</h2>
        {"<div class='msg-err'>"+msg+"</div>" if msg else ""}
        <p style="color:#aaa;margin-bottom:14px">Nombre actual: <b style="color:white">{nombre}</b></p>
        <form method="post">
            <div class="form-row">
                <label>Nuevo nombre</label>
                <input type="text" name="nuevo_nombre" value="{nombre}" style="width:100%">
            </div>
            <br>
            <button class="btn btn-green" type="submit">💾 Guardar</button>
            <a href="/admin/cargar" class="btn btn-blue" style="margin-left:8px">Cancelar</a>
        </form>
    </div>"""
    return base(content, "cargar", admin=True)

@app.route("/admin/jugador/<nombre>", methods=["GET","POST"])
def admin_ver_jugador(nombre):
    r = require_admin()
    if r: return r
    s = load_state()
    if nombre not in s["jugadores"]:
        return redirect("/admin/cargar")

    # Guardar edición de goleador
    msg_gol = ""
    if request.method == "POST":
        nuevo_gol = request.form.get("goleador", "").strip()
        s["jugadores"][nombre]["goleador"] = nuevo_gol
        save_state(s)
        msg_gol = "✅ Goleador actualizado"

    jugador = s["jugadores"][nombre]
    np = s.get("nombres_partidos", {})
    filas = ""
    for f, pred in jugador["partidos"].items():
        of = s["oficiales"].get(f"p{f}", "")
        nombre_partido = np.get(str(f), f"Partido {f}")
        if not of:
            estado = '<span style="color:#aaa">⏳ Pendiente</span>'
        else:
            r_of = parse_resultado(of)
            r_p  = parse_resultado(pred)
            if not r_p:
                estado = '<span style="color:#aaa">— Sin pronóstico</span>'
            elif not r_of:
                estado = '<span style="color:#aaa">⏳ Pendiente</span>'
            elif r_of[1]==r_p[1] and r_of[2]==r_p[2]:
                estado = '<span style="color:#4adf7a">✅ Exacto</span>'
            elif r_of[0]==r_p[0] and (r_of[1]==r_p[1] or r_of[2]==r_p[2]):
                estado = '<span style="color:#a0df4a">🟡 Un gol</span>'
            elif r_of[0]==r_p[0]:
                estado = '<span style="color:#dfdf4a">🟡 Signo</span>'
            elif r_of[1]==r_p[1] or r_of[2]==r_p[2]:
                estado = '<span style="color:#df944a">🟠 Un gol (falla)</span>'
            else:
                estado = '<span style="color:#df4a4a">❌ Fallo</span>'
        pred_show = pred if pred and pred not in ("-","0|-","") else "—"
        of_show   = of   if of else "—"
        filas += f"""<tr>
            <td style="text-align:left;color:#ddd;font-size:0.85em">{nombre_partido}</td>
            <td>{pred_show}</td>
            <td>{of_show}</td>
            <td>{estado}</td>
        </tr>"""
    of_d16  = {normalizar(v) for k,v in s["oficiales"].items() if k.startswith("d16_") and v}
    of_oct  = {normalizar(v) for k,v in s["oficiales"].items() if k.startswith("oct_") and v}
    of_cua  = {normalizar(v) for k,v in s["oficiales"].items() if k.startswith("cua_") and v}
    of_sem  = {normalizar(v) for k,v in s["oficiales"].items() if k.startswith("sem_") and v}
    of_fin  = {normalizar(v) for k,v in s["oficiales"].items() if k.startswith("fin_") and v}
    of_cam  = normalizar(s["oficiales"].get("campeon", ""))

    def elim_filas(key, of_set, label_fase):
        rows = ""
        for pred in jugador[key].values():
            if not pred or pred in ("-", ""): continue
            pred_n = normalizar(pred)
            if not of_set:
                estado = "<span style='color:#aaa'>⏳ Pendiente</span>"
            elif pred_n in of_set:
                estado = "<span style='color:#4adf7a'>✅ Acierto</span>"
            else:
                estado = "<span style='color:#df4a4a'>❌ Fallo</span>"
            rows += "<tr><td style='text-align:left;color:#ddd;font-size:0.85em'>" + label_fase + "</td><td>" + pred + "</td><td>" + estado + "</td></tr>"
        return rows

    filas_elim = (
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Dieciseisavos</td></tr>" + elim_filas("dieciseisavos", of_d16, "Clasificado") +
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Octavos</td></tr>" + elim_filas("octavos", of_oct, "Clasificado") +
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Cuartos</td></tr>" + elim_filas("cuartos", of_cua, "Clasificado") +
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Semifinales</td></tr>" + elim_filas("semis", of_sem, "Clasificado") +
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Finalistas</td></tr>" + elim_filas("finalistas", of_fin, "Clasificado")
    )

    cam_pred = jugador.get("campeon", "") or "—"
    gol_pred = jugador.get("goleador", "") or "—"
    if not of_cam:
        cam_estado = "<span style='color:#aaa'>⏳ Pendiente</span>"
    elif normalizar(cam_pred) == of_cam:
        cam_estado = "<span style='color:#4adf7a'>✅ Acierto</span>"
    else:
        cam_estado = "<span style='color:#df4a4a'>❌ Fallo</span>"

    filas_honor = (
        "<tr><td colspan='3' style='color:#4a9eff;font-weight:600;padding-top:12px'>Cuadro de Honor</td></tr>"
        "<tr><td style='text-align:left;color:#ddd;font-size:0.85em'>🥇 Campeón</td><td>" + cam_pred + "</td><td>" + cam_estado + "</td></tr>"
        "<tr><td style='text-align:left;color:#ddd;font-size:0.85em'>👟 Bota de Oro</td>""<td colspan='2'>""<form method='post' style='display:flex;gap:8px;align-items:center;flex-wrap:wrap'>""<span style='color:#aaa;font-size:0.8em'>Original: <i>" + gol_pred + "</i></span>""<input type='text' name='goleador' value='" + gol_pred + "' style='width:160px;padding:4px 8px'>""<button class='btn btn-green' type='submit' style='padding:4px 10px;font-size:0.8em'>💾 Guardar</button>""</form></td></tr>"
    )

    det = calcular_puntos(jugador, s["oficiales"], s["baremo"])
    content = (
        "<div class='card'>"
        + ("<div class='msg-ok'>" + msg_gol + "</div>" if msg_gol else "") +
        "<h2>👁 Pronósticos de " + nombre + "</h2>"
        "<p style='color:#aaa;margin-bottom:14px'>"
        "Total: <b style='color:#ffd700'>" + str(det["total"]) + " pts</b> &nbsp;|&nbsp; "
        "Exactos: <b style='color:#4adf7a'>" + str(det["exactos"]) + "</b> &nbsp;|&nbsp; "
        "Campeón: <b style='color:#4a9eff'>" + cam_pred + "</b>"
        "</p>"
        "<details open><summary style='color:#4a9eff;font-weight:600;padding:8px 0;cursor:pointer'>▶ Partidos de Grupo</summary>"
        "<div class='tabla-wrap'><table class='tabla' style='min-width:400px'>"
        "<thead><tr><th style='text-align:left'>Partido</th><th>Pronóstico</th><th>Oficial</th><th>Estado</th></tr></thead>"
        "<tbody>" + filas + "</tbody></table></div></details>"
        "<hr class='sep'>"
        "<details><summary style='color:#4a9eff;font-weight:600;padding:8px 0;cursor:pointer'>▶ Eliminatorias y Cuadro de Honor</summary>"
        "<div class='tabla-wrap'><table class='tabla' style='min-width:300px'>"
        "<thead><tr><th style='text-align:left'>Fase</th><th>Pronóstico</th><th>Estado</th></tr></thead>"
        "<tbody>" + filas_elim + filas_honor + "</tbody></table></div></details>"
        "<br><a href='/admin/cargar' class='btn btn-blue'>← Volver</a>"
        "</div>"
    )
    return base(content, "cargar", admin=True)


if __name__ == "__main__":
    import webbrowser
    webbrowser.open("http://localhost:5000")
    app.run(debug=True, port=5000)